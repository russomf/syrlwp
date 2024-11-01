# forecast.py
# Streamlining Your Research Laboratory with Python
# Authors:   Mark F. Russo, Ph.D and William Neil 
# Publisher: John Wiley & Sons, Inc.
# License:   MIT (https://opensource.org/licenses/MIT)

# Forecast of sample processing throughput using a Monte Carlo method
# given historical data.
import openpyxl, random

# Load an existing workbook from a file
wb = openpyxl.load_workbook('forecast.xlsx')

# Get throughput sample counts Data from worksheet.
ws_daily = wb['Data']

daily = []                              # List to collect daily throughput
row = 1                                 # Count rows until no data remaining
while (nsamples := ws_daily.cell(row, 1).value) != None:
    daily.append( nsamples )
    row += 1

# Perform Monte Carlo simulation to generate forecast
histogram = [0]*(15*4*7)                # Max is 15/day*7 days/week*4 weeks

# Perform 1000000 trial simulations
for i in range(1000000):                # Trials to run for simulation
    forecast = [0]*(4*7)                # Forecast over a 4 week time period
    for j in range(len(forecast)):      # Set daily throughput for four weeks
        idx = random.randint(0, len(daily)-1)
        forecast[j] = daily[idx]        # Randomly select throughput from data.
    total = sum(forecast)               # Month sum for this trial
    histogram[total] += 1               # Accumulate histogram counts

# Insert a new results worksheet with default name and set column titles
ws_results = wb.create_sheet()
ws_results.cell(1, 1).value = 'Count'
ws_results.cell(1, 2).value = 'Frequency'
ws_results.cell(1, 3).value = 'Probability'

# Compute normalized cumulative sum in [0.0, 1.0].
row = 2                                 # Skip column title row
cumulative_sum = 0                      # To compute cumulative sum
total_count    = sum(histogram)         # Total count for cumulative sum
for i in range(300, 49, -1):            # Accumulate from max to min
    cumulative_sum += histogram[i]/total_count
    ws_results.cell(row, 1).value = i   # Write results to spreadsheet
    ws_results.cell(row, 2).value = histogram[i]
    ws_results.cell(row, 3).value = cumulative_sum
    row += 1

wb.save('forecast.xlsx')                # Save update to file
wb.close()                              # Close the workbook
